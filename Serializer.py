"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ Fardan Apex --- Serializer ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
This application is responsible for serializing order items in the Fardan Apex system.

Author: Behnam Rabieyan
Company: Garma Gostar Fardan
Created: 2025
"""

# Standard library imports
import getpass
import json
import logging
import os
import re
import sys
from datetime import datetime, date
from functools import wraps

# Third-party library imports
# Make sure to install jdatetime: pip install jdatetime
import jdatetime
from cryptography.fernet import Fernet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# PyQt5 imports
from PyQt5.QtCore import Qt, QTimer, pyqtSignal
from PyQt5.QtGui import (
    QFont, QFontDatabase, QIcon, QIntValidator, QPixmap, QTextOption
)
from PyQt5.QtWidgets import (
    QApplication, QComboBox, QDialog, QFileDialog, QFormLayout, QGroupBox,
    QHeaderView, QHBoxLayout, QInputDialog, QLabel, QLineEdit, QListWidget,
    QListWidgetItem, QMainWindow, QMessageBox, QProgressBar, QProgressDialog,
    QPushButton, QSplashScreen, QTabWidget, QTableWidget, QTableWidgetItem,
    QTextEdit, QVBoxLayout, QWidget
)

# ---------- تنظیمات ----------
logging.basicConfig(
    level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s'
)
ADMIN_PASSWORD = "Ggfco@2020"

SECRET_KEY = b"SnZFJqzdj1xx6rxksdPL5P_-UKijvx4DRlR0a5-s1lQ="
cipher = Fernet(SECRET_KEY)

# مقادیر پیش‌فرض که از فایل تنظیمات خوانده می‌شوند
EXCEL_FILE = r"\\fileserver\Mohandesi\Serializer\orders.xlsx"
SHEET_NAME = "OrderList"
TABLE_NAME = "ordertable"
ALLOWED_USERS = []

# ---------- توابع کمکی برای مسیرها و فایل‌ها ----------
def resource_path(relative_path: str) -> str:
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def app_dir_path(relative_path: str) -> str:
    """ Get absolute path in the app's directory """
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, relative_path)


SETTINGS_FILE = app_dir_path("settings.json")
CACHE_FILE = app_dir_path("excel_data.cache")

PRODUCT_MAP = {
    "MF": "F", "MR": "R", "MU": "U", "نفراست": "ن", "فویلی": "ف",
    "هیتر سیمی": "س", "لوله ای دیفراست": "د", "ترموسوییچ": "TS",
    "ترموفیوز": "TF"
}
GROUP_M = {"MF", "MR", "MU"}


# ---------- توابع کاربردی (Helpers) ----------
def shamsi_to_gregorian_str(shamsi_str: str) -> str | None:
    """
    رشته تاریخ شمسی با فرمت YYYY-MM-DD را به رشته تاریخ میلادی تبدیل می‌کند.
    """
    if not shamsi_str or not re.match(r"^\d{4}-\d{2}-\d{2}$", shamsi_str):
        return None
    try:
        year, month, day = map(int, shamsi_str.split('-'))
        shamsi_date = jdatetime.date(year, month, day)
        gregorian_date = shamsi_date.togregorian()
        return gregorian_date.strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        return None

def gregorian_to_shamsi_str(gregorian_input) -> str:
    """
    تاریخ میلادی (آبجکت دیت‌تایم یا رشته) را به رشته تاریخ شمسی تبدیل می‌کند.
    این تابع در برابر فرمت‌های مختلف و خطاهای احتمالی مقاوم است.
    """
    if not gregorian_input:
        return ""

    date_obj = None

    if isinstance(gregorian_input, (datetime, date)):
        date_obj = gregorian_input
    elif isinstance(gregorian_input, str):
        date_str = gregorian_input.split(' ')[0]
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            return "" 

    if date_obj:
        try:
            return jdatetime.date.fromgregorian(date=date_obj).strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            return ""
    
    return ""

def initialize_settings():
    """این تابع تنظیمات را از فایل خوانده و متغیرهای گلوبال را مقداردهی می‌کند"""
    global EXCEL_FILE, SHEET_NAME, TABLE_NAME, ALLOWED_USERS
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "rb") as f:
                encrypted = f.read()
                if not encrypted:
                    return
                raw = cipher.decrypt(encrypted)
                settings = json.loads(raw.decode("utf-8"))
                EXCEL_FILE = settings.get("excel_file", EXCEL_FILE)
                SHEET_NAME = settings.get("sheet_name", SHEET_NAME)
                TABLE_NAME = settings.get("table_name", TABLE_NAME)
                ALLOWED_USERS = settings.get("allowed_users", ALLOWED_USERS)
    except Exception as e:
        logging.error(f"Error loading settings: {e}")
        QMessageBox.warning(
            None, "خطا در بارگذاری تنظیمات",
            f"فایل تنظیمات قابل خواندن نیست. مقادیر پیش‌فرض استفاده خواهد شد.\n{e}"
        )


def save_settings(data: dict):
    """ Encrypts and saves settings to the settings file. """
    try:
        raw = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        encrypted = cipher.encrypt(raw)
        with open(SETTINGS_FILE, "wb") as f:
            f.write(encrypted)
        return True
    except Exception as e:
        logging.error(f"Error saving settings: {e}")
        return False


def update_excel_table_range(ws, table_name):
    """ Updates the range of an Excel table to include all rows. """
    try:
        table = ws.tables[table_name]
        start_cell, end_cell = table.ref.split(':')
        min_col, min_row = ws[start_cell].col_idx, ws[start_cell].row
        max_col = ws[end_cell].col_idx
        max_row = ws.max_row
        new_ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )
        table.ref = new_ref
    except KeyError:
        logging.warning(
            f"Table '{table_name}' not found. Data saved but table not updated."
        )


def ensure_excel(show_message=True) -> bool:
    """ Checks if the Excel file exists and shows a warning if not. """
    if not os.path.exists(EXCEL_FILE):
        if show_message:
            QMessageBox.warning(None, "هشدار", f"فایل اکسل یافت نشد:\n{EXCEL_FILE}")
        return False
    return True


def normalize_farsi(text: str) -> str:
    """ Normalizes Farsi characters and strips extra whitespace. """
    if not text:
        return ""
    replacements = {"ي": "ی", "ك": "ک", "ة": "ه", "ۀ": "ه"}
    persian_to_english = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
    text = text.translate(persian_to_english)
    
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return re.sub(r"\s+", " ", text).strip()


def get_product_abbreviation(product_type: str) -> str:
    """ Returns the abbreviation for a given product type. """
    p = normalize_farsi(product_type)
    key = p.upper() if re.match(r"^[A-Za-z]{1,4}$", p) else p
    return PRODUCT_MAP.get(key, "0")


def is_in_m_group(product_type: str) -> bool:
    """ Checks if a product type belongs to the 'M' group. """
    p_norm = normalize_farsi(product_type)
    key = p_norm.upper() if re.match(r"^[A-Za-z]{1,4}$", p_norm) else p_norm
    return str(key).upper() in GROUP_M


def next_item_and_serial(date_text, product_type, max_groupA, max_groupB):
    """ Generates the next item index and serial number for a product. """
    abbrev = get_product_abbreviation(product_type)
    shamsi_year_match = re.search(r"^\d{4}", date_text)
    shamsi_year = shamsi_year_match.group(0) if date_text and shamsi_year_match else "0000"

    if is_in_m_group(product_type):
        max_groupA += 1
        item_idx = max_groupA
    else:
        max_groupB += 1
        item_idx = max_groupB

    serial = f"{item_idx}-{shamsi_year}-{abbrev}"
    return item_idx, serial, max_groupA, max_groupB


class ExcelDataManager:
    """ Manages reading from and writing to the Excel file, with caching. """
    def __init__(self, excel_path, sheet_name, table_name):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.table_name = table_name
        self._cache = None

    def _is_cache_valid(self):
        if not os.path.exists(CACHE_FILE) or not os.path.exists(self.excel_path):
            return False
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
            excel_mtime = os.path.getmtime(self.excel_path)
            if cached_data.get("mtime") == excel_mtime:
                search_index = cached_data.get("search_index", {})
                for order_no, rows in search_index.items():
                    for row in rows:
                        if len(row) > 1 and isinstance(row[1], str) and row[1]:
                            try:
                                row[1] = datetime.fromisoformat(row[1])
                            except ValueError:
                                pass
                cached_data["search_index"] = search_index
                self._cache = cached_data
                return True
        except (json.JSONDecodeError, KeyError, FileNotFoundError):
            return False
        return False

    def _build_cache(self):
        logging.info("Cache invalid or missing. Rebuilding from Excel file...")
        if not ensure_excel(show_message=False):
            return None
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            ws = wb[self.sheet_name]
            maxA, maxB, max_rowid = self._compute_maxes_from_ws(ws)
            
            search_index_with_objects = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                order_no = str(row[2])
                if order_no not in search_index_with_objects:
                    search_index_with_objects[order_no] = []
                
                row_data = []
                for i, cell in enumerate(row):
                    if i == 1 and isinstance(cell, datetime):
                         row_data.append(cell)
                    else:
                         row_data.append(str(cell) if cell is not None else "")
                search_index_with_objects[order_no].append(row_data)

            wb.close()
            excel_mtime = os.path.getmtime(self.excel_path)
            
            serializable_search_index = {}
            for key, rows in search_index_with_objects.items():
                serializable_rows = [
                    [item.isoformat() if isinstance(item, (datetime, date)) else item for item in row]
                    for row in rows
                ]
                serializable_search_index[key] = serializable_rows
            
            self._cache = {
                "mtime": excel_mtime, "maxA": maxA, "maxB": maxB,
                "max_rowid": max_rowid, "search_index": serializable_search_index
            }
            with open(CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(self._cache, f)
            
            self._cache['search_index'] = search_index_with_objects
            logging.info("Cache rebuild complete.")
            return self._cache
        except Exception as e:
            QMessageBox.critical(
                None, "خطای ساخت کش",
                f"امکان خواندن فایل اکسل برای ساخت کش وجود نداشت.\n{e}"
            )
            return None

    def _get_data(self):
        if self._cache and self._is_cache_valid():
            return self._cache
        if self._is_cache_valid():
            return self._cache
        return self._build_cache()

    def _compute_maxes_from_ws(self, ws):
        maxA, maxB, max_rowid = 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                rowid = int(row[0]) if row[0] is not None else 0
            except (ValueError, TypeError): rowid = 0
            if rowid > max_rowid: max_rowid = rowid

            ptype = normalize_farsi(str(row[3] or ""))
            try:
                item_idx = int(row[6]) if row[6] is not None else 0
            except (ValueError, TypeError): item_idx = 0

            if is_in_m_group(ptype):
                if item_idx > maxA: maxA = item_idx
            else:
                if item_idx > maxB: maxB = item_idx
        return maxA, maxB, max_rowid

    def get_order(self, order_no):
        data = self._get_data()
        if data and "search_index" in data:
            return data["search_index"].get(str(order_no), [])
        return []

    def get_max_values(self):
        data = self._get_data()
        if data:
            return (data.get("maxA", 0), data.get("maxB", 0), data.get("max_rowid", 0))
        try:
            logging.warning("Cache failed, reading max values directly from Excel.")
            wb = load_workbook(self.excel_path, read_only=True)
            ws = wb[self.sheet_name]
            results = self._compute_maxes_from_ws(ws)
            wb.close()
            return results
        except Exception:
            return 0, 0, 0

    def invalidate_cache(self):
        if os.path.exists(CACHE_FILE):
            try:
                os.remove(CACHE_FILE)
                self._cache = None
                logging.info("Cache invalidated due to data change.")
            except OSError as e:
                logging.error(f"Could not delete cache file: {e}")

APP_STYLESHEET = """
QWidget { background: #f5f7fb; font-family: 'Segoe UI', Tahoma, Arial; }
QLineEdit, QTextEdit, QComboBox {
    background: white; border: 1px solid #d0d7df; border-radius: 6px; padding: 6px;
}
QPushButton {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #5aa9ff, stop:1 #2e7dff);
    color: white; border: none; padding: 8px 12px; border-radius: 8px;
}
QPushButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #6bb8ff, stop:1 #3b8bff);
}
QPushButton:disabled { background-color: #d0d7df; color: #8c959f; }
QPushButton#secondary { background: #eef4ff; color: #1a3b6e; border: 1px solid #d0dbff; }
QTableWidget { background: white; border: 1px solid #e0e6ef; gridline-color: #f1f5fb; }
QHeaderView::section { background: #eef4ff; padding: 6px; border: none; }
QTabWidget::pane { border: none; }
QTabBar::tab { background: transparent; padding: 8px 16px; }
"""

class ProductDialog(QDialog):
    """ Dialog for adding or editing a product entry. """
    product_added = pyqtSignal(tuple)

    def __init__(self, parent=None, preset=None):
        super().__init__(parent)
        self.setWindowTitle("افزودن/ویرایش محصول")
        self.setFixedSize(460, 220)
        self.setFont(QFont("Segoe UI", 10))
        form = QFormLayout(self)
        form.setLabelAlignment(Qt.AlignRight)
        self.cb_type = QComboBox()
        self.cb_type.addItems(
            [
            '', 'فویلی', 'هیتر سیمی', 'نفراست', 'لوله ای دیفراست',
            'ترموفیوز', 'ترموسوییچ', 'لوله استیل قطر 7 (60میل)',
            'MF', 'MR', 'MU'
            ]
        )
        self.cb_type.setEditable(True)
        self.e_code = QLineEdit()
        self.e_code.setAlignment(Qt.AlignRight)
        self.e_qty = QLineEdit()
        self.e_qty.setValidator(QIntValidator(1, 10000000, self))
        form.addRow("نوع محصول ", self.cb_type)
        form.addRow("کد محصول ", self.e_code)
        form.addRow("تعداد ", self.e_qty)
        if preset:
            self.cb_type.setCurrentText(preset[0])
            self.e_code.setText(preset[1])
            self.e_qty.setText(str(preset[2]))
        btn_layout = QHBoxLayout()
        btn_register = QPushButton("ثبت")
        btn_close = QPushButton("بستن")
        btn_close.setObjectName("secondary")
        btn_layout.addStretch()
        btn_layout.addWidget(btn_register)
        btn_layout.addWidget(btn_close)
        form.addRow(btn_layout)
        btn_register.clicked.connect(self.on_register)
        btn_close.clicked.connect(self.reject)

    def on_register(self):
        ptype = normalize_farsi(self.cb_type.currentText())
        code = normalize_farsi(self.e_code.text())
        qty_text = self.e_qty.text()
        if not all([ptype, code, qty_text]):
            QMessageBox.critical(self, "خطا", "همه فیلدها الزامی هستند.")
            return
        try:
            qty = int(qty_text)
            assert qty > 0
        except (ValueError, AssertionError):
            QMessageBox.critical(self, "خطا", "تعداد نامعتبر است.")
            return
        self.product_added.emit((ptype, code, qty))
        self.cb_type.setCurrentIndex(0)
        self.e_code.clear()
        self.e_qty.clear()
        self.cb_type.setFocus()

def with_progress_dialog(title, label):
    def decorator(func):
        @wraps(func)
        def wrapper(self, *args, **kwargs):
            progress = QProgressDialog(label, "لغو", 0, 0, self)
            progress.setWindowTitle(title)
            progress.setWindowModality(Qt.ApplicationModal)
            progress.setMinimumDuration(0)
            progress.show()
            QApplication.processEvents()
            try:
                return func(self, *args, **kwargs)
            finally:
                progress.close()
        return wrapper
    return decorator

class App(QMainWindow):
    """ Main application window. """
    def __init__(self, is_authorized: bool):
        super().__init__()
        self.is_authorized = is_authorized
        self.data_manager = ExcelDataManager(EXCEL_FILE, SHEET_NAME, TABLE_NAME)
        self.setWindowTitle("Serializer - Production Serial Generator")
        self.resize(900, 500)
        self.setStyleSheet(APP_STYLESHEET)
        self.statusBar()
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        tab_style = """
        QTabBar::tab:selected { background: #ffffff; color: #111827; font-weight: bold; border-bottom: 3px solid #2563eb; }
        QTabBar::tab:!selected { background: #e5e7eb; }
        QTabWidget::pane { border-top: 2px solid #d1d5db; background: #ffffff; }
        """
        self.tabs.setStyleSheet(tab_style)
        main_layout.addWidget(self.tabs)
        self.tab_new = QWidget()
        self.tab_search = QWidget()
        self.tab_option = QWidget()
        self.tabs.addTab(self.tab_new, "ثبت سفارش جدید")
        self.tabs.addTab(self.tab_search, "جستجو و ویرایش")
        self.tabs.addTab(self.tab_option, "ویژگی‌ها")
        self.build_tab_new()
        self.build_tab_search()
        self.build_tab_option()
        self.apply_access_control()

    def apply_access_control(self):
        if self.is_authorized:
            self.statusBar().showMessage("حساب کاربری شما فعال است", 5000)
            return
        self.statusBar().showMessage("حالت مهمان: شما مجوز ایجاد تغییرات را ندارید.")
        for btn in [self.btn_save_new, self.btn_add_new, self.btn_edit_new, self.btn_del_new,
                    self.btn_save_search, self.btn_add_search, self.btn_edit_search,
                    self.btn_save_options, self.btn_add_user, self.btn_remove_user, self.btn_browse_excel]:
            btn.setEnabled(False)
        self.user_input.setEnabled(False)
        self.e_excel_path.setReadOnly(True)
        self.e_sheet_name.setReadOnly(True)
        self.e_table_name.setReadOnly(True)

    def _create_order_tab_widgets(self, is_search_tab=False):
        top_layout = QHBoxLayout()
        order_no = QLineEdit()
        order_no.setAlignment(Qt.AlignRight)
        order_no.setFixedWidth(220)
        date = QLineEdit()
        date.setAlignment(Qt.AlignRight)
        date.setFixedWidth(160)
        date.setPlaceholderText("مثال: ۱۴۰۳۰۶۰۹")
        top_layout.addWidget(QLabel("شماره سفارش "), 0, Qt.AlignRight)
        top_layout.addWidget(order_no)
        top_layout.addSpacing(12)
        top_layout.addWidget(QLabel("تاریخ "), 0, Qt.AlignRight)
        top_layout.addWidget(date)
        desc_layout = QHBoxLayout()
        description = QLineEdit()
        description.setAlignment(Qt.AlignRight)
        desc_layout.addWidget(QLabel("توضیحات "), 0, Qt.AlignRight)
        desc_layout.addWidget(description)
        cols = 4 if is_search_tab else 3
        headers = ["حذف", "نوع محصول", "کد محصول", "تعداد"] if is_search_tab else ["نوع محصول", "کد محصول", "تعداد"]
        table = QTableWidget(0, cols)
        table.setHorizontalHeaderLabels(headers)
        header = table.horizontalHeader()
        if is_search_tab:
            header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        start_col = 1 if is_search_tab else 0
        for i in range(start_col, cols):
            header.setSectionResizeMode(i, QHeaderView.Stretch)
        table.verticalHeader().setVisible(False)
        serial_box = QTextEdit()
        serial_box.setReadOnly(True)
        serial_box.setWordWrapMode(QTextOption.NoWrap)
        serial_box.setLayoutDirection(Qt.LeftToRight)
        serial_box.setFont(QFont("Consolas", 10))
        return {"top_layout": top_layout, "order_no": order_no, "date": date, "desc_layout": desc_layout,
                "description": description, "table": table, "serial_box": serial_box}

    def _add_product_to_table(self, table: QTableWidget):
        dlg = ProductDialog(self)
        def add_row(data):
            rc = table.rowCount()
            table.insertRow(rc)
            new_item_marker = QTableWidgetItem(data[0])
            new_item_marker.setData(Qt.UserRole, None)
            if table.columnCount() == 4:
                btn_del = QPushButton("×")
                btn_del.setStyleSheet("color: red; font-weight: bold;")
                btn_del.clicked.connect(self.toggle_row_for_deletion)
                table.setCellWidget(rc, 0, btn_del)
                table.setItem(rc, 1, new_item_marker)
                table.setItem(rc, 2, QTableWidgetItem(data[1]))
                table.setItem(rc, 3, QTableWidgetItem(str(data[2])))
            else:
                for i, val in enumerate(data):
                    table.setItem(rc, i, QTableWidgetItem(str(val)))
        dlg.product_added.connect(add_row)
        dlg.exec_()

    def _edit_product_in_table(self, table: QTableWidget):
        row = table.currentRow()
        if row == -1: return
        start_col = 1 if table.columnCount() == 4 else 0
        preset_data = [table.item(row, i).text() for i in range(start_col, table.columnCount())]
        dlg = ProductDialog(self, preset=preset_data)
        def update_row(data):
            for i, val in enumerate(data):
                table.setItem(row, i + start_col, QTableWidgetItem(str(val)))
            dlg.accept()
        dlg.product_added.connect(update_row)
        dlg.exec_()

    def _copy_serials_to_clipboard(self, serial_box: QTextEdit):
        if not serial_box.toPlainText().strip(): return
        QApplication.clipboard().setText(serial_box.toPlainText())
        QMessageBox.information(self, "کپی شد", "سریال‌ها به کلیپ‌بورد کپی شدند.")

    def build_tab_new(self):
        main_hbox = QHBoxLayout(self.tab_new)
        left_layout, right_layout = QVBoxLayout(), QVBoxLayout()
        main_hbox.addLayout(left_layout, 3)
        main_hbox.addLayout(right_layout, 1)
        widgets = self._create_order_tab_widgets()
        self.new_order_no = widgets["order_no"]
        self.new_date = widgets["date"]
        self.new_desc = widgets["description"]
        self.table_new = widgets["table"]
        self.serial_box = widgets["serial_box"]
        self.new_date.setMaxLength(10)
        self.new_date.textChanged.connect(self.format_date_input)
        btn_new_top = QPushButton("ثبت سفارش جدید")
        btn_new_top.setObjectName("secondary")
        btn_new_top.clicked.connect(self.reset_new_order_form)
        widgets["top_layout"].addWidget(btn_new_top)
        widgets["top_layout"].addStretch()
        left_layout.addLayout(widgets["top_layout"])
        left_layout.addLayout(widgets["desc_layout"])
        left_layout.addWidget(self.table_new)
        btn_layout = QHBoxLayout()
        self.btn_add_new = QPushButton("افزودن محصول")
        self.btn_edit_new = QPushButton("ویرایش محصول")
        self.btn_del_new = QPushButton("حذف محصول")
        self.btn_save_new = QPushButton("ذخیره سفارش")
        btn_layout.addWidget(self.btn_add_new)
        btn_layout.addWidget(self.btn_edit_new)
        btn_layout.addWidget(self.btn_del_new)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_save_new)
        left_layout.addLayout(btn_layout)
        right_layout.addWidget(QLabel("سریال‌های این سفارش:"))
        right_layout.addWidget(self.serial_box)
        btn_copy = QPushButton("کپی سریال‌ها")
        btn_copy.setFixedWidth(130)
        right_layout.addWidget(btn_copy)
        self.btn_add_new.clicked.connect(lambda: self._add_product_to_table(self.table_new))
        self.btn_edit_new.clicked.connect(lambda: self._edit_product_in_table(self.table_new))
        self.btn_del_new.clicked.connect(lambda: self.delete_selected(self.table_new))
        self.btn_save_new.clicked.connect(self.save_order_new)
        btn_copy.clicked.connect(lambda: self._copy_serials_to_clipboard(self.serial_box))

    def format_date_input(self, text):
        sender = self.sender()
        if not sender: return
        sender.blockSignals(True)
        cleaned_text = normalize_farsi(text)
        cleaned_text = re.sub(r'\D', '', cleaned_text)
        if len(cleaned_text) > 8: cleaned_text = cleaned_text[:8]
        formatted_text = ""
        if len(cleaned_text) > 0: formatted_text = cleaned_text[:4]
        if len(cleaned_text) > 4: formatted_text += '-' + cleaned_text[4:6]
        if len(cleaned_text) > 6: formatted_text += '-' + cleaned_text[6:]
        sender.setText(formatted_text)
        sender.setCursorPosition(len(formatted_text))
        sender.blockSignals(False)

    def build_tab_search(self):
        main_hbox = QHBoxLayout(self.tab_search)
        left_layout, right_layout = QVBoxLayout(), QVBoxLayout()
        main_hbox.addLayout(left_layout, 3)
        main_hbox.addLayout(right_layout, 1)
        widgets = self._create_order_tab_widgets(is_search_tab=True)
        self.search_order_no = widgets["order_no"]
        self.search_date = widgets["date"]
        self.search_desc = widgets["description"]
        self.table_search = widgets["table"]
        self.serial_box_search = widgets["serial_box"]
        self.search_date.textChanged.connect(self.format_date_input)
        btn_search = QPushButton("جستجو")
        widgets["top_layout"].insertWidget(2, btn_search)
        widgets["top_layout"].addStretch()
        left_layout.addLayout(widgets["top_layout"])
        left_layout.addLayout(widgets["desc_layout"])
        left_layout.addWidget(self.table_search)
        btn_layout = QHBoxLayout()
        self.btn_add_search = QPushButton("افزودن محصول")
        self.btn_edit_search = QPushButton("ویرایش محصول")
        self.btn_save_search = QPushButton("ذخیره تغییرات")
        btn_layout.addWidget(self.btn_add_search)
        btn_layout.addWidget(self.btn_edit_search)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_save_search)
        left_layout.addLayout(btn_layout)
        right_layout.addWidget(QLabel("سریال‌های این سفارش:"))
        right_layout.addWidget(self.serial_box_search)
        btn_copy = QPushButton("کپی سریال‌ها")
        btn_copy.setFixedWidth(130)
        right_layout.addWidget(btn_copy)
        btn_search.clicked.connect(self.search_order)
        self.btn_add_search.clicked.connect(lambda: self._add_product_to_table(self.table_search))
        self.btn_edit_search.clicked.connect(lambda: self._edit_product_in_table(self.table_search))
        self.btn_save_search.clicked.connect(self.save_changes_search)
        btn_copy.clicked.connect(lambda: self._copy_serials_to_clipboard(self.serial_box_search))

    def build_tab_option(self):
        main_layout = QVBoxLayout(self.tab_option)
        h_main = QHBoxLayout()
        main_layout.addLayout(h_main)
        self.user_group = QGroupBox("کاربران مجاز")
        self.user_group.setFixedWidth(280)
        user_group_layout = QVBoxLayout(self.user_group)
        h_main.addWidget(self.user_group, 0, Qt.AlignTop)
        user_top_row = QHBoxLayout()
        self.btn_add_user = QPushButton("افزودن")
        self.btn_add_user.setFixedWidth(80)
        self.user_input = QLineEdit()
        self.user_input.setPlaceholderText("نام کاربری جدید...")
        user_top_row.addWidget(self.user_input)
        user_top_row.addWidget(self.btn_add_user)
        user_group_layout.addLayout(user_top_row)
        self.user_list = QListWidget()
        self.user_list.setLayoutDirection(Qt.LeftToRight)
        user_group_layout.addWidget(self.user_list)
        self.btn_remove_user = QPushButton("حذف کاربر انتخاب شده")
        self.btn_remove_user.setObjectName("secondary")
        user_group_layout.addWidget(self.btn_remove_user)
        for u in ALLOWED_USERS: self.add_user_item(u)
        group_settings = QGroupBox("تنظیمات اکسل")
        settings_layout = QFormLayout(group_settings)
        h_main.addWidget(group_settings, 1, Qt.AlignTop)
        file_row = QHBoxLayout()
        self.btn_browse_excel = QPushButton("انتخاب فایل")
        self.btn_browse_excel.setFixedWidth(120)
        self.e_excel_path = QLineEdit(EXCEL_FILE)
        file_row.addWidget(self.e_excel_path)
        file_row.addWidget(self.btn_browse_excel)
        settings_layout.addRow("آدرس فایل اکسل:", file_row)
        self.e_sheet_name = QLineEdit(SHEET_NAME)
        settings_layout.addRow("نام برگه:", self.e_sheet_name)
        self.e_table_name = QLineEdit(TABLE_NAME)
        settings_layout.addRow("نام جدول:", self.e_table_name)
        bottom_row = QHBoxLayout()
        btn_about = QPushButton("درباره برنامه")
        self.btn_save_options = QPushButton("ذخیره")
        bottom_row.addWidget(btn_about)
        bottom_row.addStretch()
        bottom_row.addWidget(self.btn_save_options)
        main_layout.addLayout(bottom_row)
        self.btn_add_user.clicked.connect(self.add_user)
        self.btn_remove_user.clicked.connect(self.remove_selected_user)
        self.btn_browse_excel.clicked.connect(self.browse_excel_file)
        btn_about.clicked.connect(self.show_about)
        self.btn_save_options.clicked.connect(self.save_options)

    @with_progress_dialog("در حال پردازش", "در حال ذخیره سفارش...")
    def save_order_new(self, checked=False):
        if not ensure_excel(): return
        shamsi_date_text = normalize_farsi(self.new_date.text())
        order_no = normalize_farsi(self.new_order_no.text())
        gregorian_date_text = shamsi_to_gregorian_str(shamsi_date_text)
        if not gregorian_date_text:
            QMessageBox.critical(self, "خطا", "تاریخ وارد شده نامعتبر است. لطفاً فرمت صحیح (YYYY-MM-DD) را رعایت کنید.")
            return
        if not order_no:
            QMessageBox.critical(self, "خطا", "شماره سفارش الزامی است.")
            return
        if self.data_manager.get_order(order_no):
            QMessageBox.warning(self, "سفارش تکراری", f"سفارش با شماره '{order_no}' قبلا ثبت شده است.")
            return
        items = []
        for r in range(self.table_new.rowCount()):
            try:
                items.append((self.table_new.item(r, 0).text(), self.table_new.item(r, 1).text(), int(self.table_new.item(r, 2).text())))
            except (AttributeError, ValueError):
                QMessageBox.critical(self, "خطا", f"داده نامعتبر در ردیف {r + 1}")
                return
        if not items:
            QMessageBox.critical(self, "خطا", "حداقل یک محصول باید اضافه شود.")
            return
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            maxA, maxB, max_rowid = self.data_manager.get_max_values()
            serial_lines, username, now_str = [], getpass.getuser(), datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for ptype, code, qty in items:
                item_idx, serial, maxA, maxB = next_item_and_serial(shamsi_date_text, ptype, maxA, maxB)
                max_rowid += 1
                ws.append([max_rowid, gregorian_date_text, order_no, ptype, code, qty, item_idx, serial, normalize_farsi(self.new_desc.text()), username, now_str])
                serial_lines.append('\u200E' + serial)
            update_excel_table_range(ws, TABLE_NAME)
            wb.save(EXCEL_FILE)
            self.data_manager.invalidate_cache()
            QMessageBox.information(self, "موفق", "سفارش با موفقیت ثبت شد.")
            self.serial_box.setPlainText("\n".join(serial_lines))
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ذخیره‌سازی: {e}")

    @with_progress_dialog("در حال پردازش", "در حال جستجوی سفارش...")
    def search_order(self, checked=False):
        order_no_to_search = ""
        try:
            order_no_to_search = normalize_farsi(self.search_order_no.text())
            if not order_no_to_search: return
            found_rows = self.data_manager.get_order(order_no_to_search)
            self.table_search.setRowCount(0)
            self.serial_box_search.clear()
            if not found_rows:
                QMessageBox.information(self, "یافت نشد", "سفارشی با این شماره پیدا نشد.")
                self.search_date.clear()
                self.search_desc.clear()
                return
            first = found_rows[0]
            gregorian_date_val = first[1] 
            shamsi_date_text = gregorian_to_shamsi_str(gregorian_date_val)
            if not shamsi_date_text and gregorian_date_val:
                self.search_date.setText(f"تاریخ نامعتبر: {gregorian_date_val}")
            else:
                self.search_date.setText(shamsi_date_text)
            self.search_desc.setText(str(first[8] or ""))
            serial_lines = []
            for r_data in found_rows:
                rc = self.table_search.rowCount()
                self.table_search.insertRow(rc)
                item_type = QTableWidgetItem(str(r_data[3] or ""))
                item_type.setData(Qt.UserRole, int(r_data[0]))
                btn_del = QPushButton("×")
                btn_del.setStyleSheet("color: red; font-weight: bold;")
                btn_del.clicked.connect(self.toggle_row_for_deletion)
                self.table_search.setCellWidget(rc, 0, btn_del)
                self.table_search.setItem(rc, 1, item_type)
                self.table_search.setItem(rc, 2, QTableWidgetItem(str(r_data[4] or "")))
                self.table_search.setItem(rc, 3, QTableWidgetItem(str(r_data[5] or "")))
                serial_lines.append('\u200E' + str(r_data[7] or ""))
            self.serial_box_search.setPlainText("\n".join(serial_lines))
        except Exception as e:
            logging.error(f"Error during order search for '{order_no_to_search}': {e}", exc_info=True)
            QMessageBox.critical(self, "خطای بحرانی", f"یک خطای پیش‌بینی نشده در هنگام جستجو رخ داد:\n\n{e}\n\nاین خطا معمولاً به دلیل وجود داده تاریخ نامعتبر در فایل اکسل برای این سفارش است.")

    def toggle_row_for_deletion(self):
        button = self.sender()
        if not button: return
        row = self.table_search.indexAt(button.pos()).row()
        item = self.table_search.item(row, 1)
        if not item: return
        font = item.font()
        is_struck_out = font.strikeOut()
        font.setStrikeOut(not is_struck_out)
        button.setText("⟲" if not is_struck_out else "×")
        button.setStyleSheet(f"color: {'green' if not is_struck_out else 'red'}; font-weight: bold;")
        for col in range(1, self.table_search.columnCount()):
            if itm := self.table_search.item(row, col):
                itm.setFont(font)

    @with_progress_dialog("در حال پردازش", "در حال ذخیره تغییرات...")
    def save_changes_search(self, checked=False):
        order_no = normalize_farsi(self.search_order_no.text())
        shamsi_date_text = normalize_farsi(self.search_date.text())
        desc_text = normalize_farsi(self.search_desc.text())
        gregorian_date_text = shamsi_to_gregorian_str(shamsi_date_text)
        if not all([order_no, gregorian_date_text]):
            QMessageBox.critical(self, "خطا", "تاریخ یا شماره سفارش نامعتبر است.")
            return
        original_order_data = self.data_manager.get_order(order_no)
        original_row_map = {int(row[0]): row for row in original_order_data}
        items_to_update, items_to_add, ids_to_delete = [], [], []
        for r in range(self.table_search.rowCount()):
            item_type = self.table_search.item(r, 1)
            if not item_type: continue
            row_id = item_type.data(Qt.UserRole)
            is_marked_for_deletion = item_type.font().strikeOut()
            item_data = {'ptype': item_type.text(), 'code': self.table_search.item(r, 2).text(), 'qty': int(self.table_search.item(r, 3).text())}
            if row_id is not None:
                if is_marked_for_deletion: ids_to_delete.append(row_id)
                else:
                    item_data['row_id'] = row_id
                    items_to_update.append(item_data)
            elif not is_marked_for_deletion:
                items_to_add.append(item_data)
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            if ids_to_delete:
                indices = [r_idx for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2) if row[0] is not None and int(row[0]) in ids_to_delete]
                for r_idx in sorted(indices, reverse=True): ws.delete_rows(r_idx, 1)
            wb.save(EXCEL_FILE)
            wb = load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            maxA, maxB, max_rowid = self.data_manager._compute_maxes_from_ws(ws)
            new_year = (re.search(r"^\d{4}", shamsi_date_text) or re.search(r"....", "0000")).group(0)
            if items_to_update:
                for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value is None: continue
                    for item in items_to_update:
                        if int(row[0].value) == item['row_id']:
                            original_row = original_row_map.get(item['row_id'])
                            original_ptype = original_row[3] if original_row else ""
                            if is_in_m_group(item['ptype']) != is_in_m_group(original_ptype):
                                new_item_idx = maxA + 1 if is_in_m_group(item['ptype']) else maxB + 1
                                if is_in_m_group(item['ptype']): maxA += 1
                                else: maxB += 1
                                new_abbrev = get_product_abbreviation(item['ptype'])
                                new_serial = f"{new_item_idx}-{new_year}-{new_abbrev}"
                                ws.cell(row=r_idx, column=7).value = new_item_idx
                                ws.cell(row=r_idx, column=8).value = new_serial
                            else:
                                if isinstance(existing_serial := ws.cell(row=r_idx, column=8).value, str) and len(parts := existing_serial.split('-')) == 3:
                                    new_abbrev = get_product_abbreviation(item['ptype'])
                                    ws.cell(row=r_idx, column=8).value = f"{parts[0]}-{new_year}-{new_abbrev}"
                            ws.cell(row=r_idx, column=2).value = gregorian_date_text
                            ws.cell(row=r_idx, column=4).value = item['ptype']
                            ws.cell(row=r_idx, column=5).value = item['code']
                            ws.cell(row=r_idx, column=6).value = item['qty']
                            ws.cell(row=r_idx, column=9).value = desc_text
                            break
            if items_to_add:
                username, now_str = getpass.getuser(), datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                for item in items_to_add:
                    item_idx, serial, maxA, maxB = next_item_and_serial(shamsi_date_text, item['ptype'], maxA, maxB)
                    max_rowid += 1
                    ws.append([max_rowid, gregorian_date_text, order_no, item['ptype'], item['code'], item['qty'], item_idx, serial, desc_text, username, now_str])
            update_excel_table_range(ws, TABLE_NAME)
            wb.save(EXCEL_FILE)
            self.data_manager.invalidate_cache()
            QMessageBox.information(self, "موفق", "تغییرات با موفقیت ذخیره شد.")
            self.search_order()
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در ذخیره‌سازی تغییرات: {e}")

    def reset_new_order_form(self):
        self.new_order_no.clear()
        self.new_date.setText("")
        self.new_desc.clear()
        self.table_new.setRowCount(0)
        self.serial_box.clear()

    def delete_selected(self, table):
        selected_rows = table.selectionModel().selectedRows()
        for idx in sorted([r.row() for r in selected_rows], reverse=True):
            table.removeRow(idx)

    def browse_excel_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "انتخاب فایل اکسل", "", "Excel Files (*.xlsx)")
        if path: self.e_excel_path.setText(path)

    def save_options(self):
        global EXCEL_FILE, SHEET_NAME, TABLE_NAME, ALLOWED_USERS
        EXCEL_FILE = self.e_excel_path.text().strip()
        SHEET_NAME = self.e_sheet_name.text().strip()
        TABLE_NAME = self.e_table_name.text().strip()
        ALLOWED_USERS = [self.user_list.item(i).text() for i in range(self.user_list.count())]
        settings = {"excel_file": EXCEL_FILE, "sheet_name": SHEET_NAME, "table_name": TABLE_NAME, "allowed_users": ALLOWED_USERS}
        if save_settings(settings):
            self.data_manager = ExcelDataManager(EXCEL_FILE, SHEET_NAME, TABLE_NAME)
            QMessageBox.information(self, "ذخیره شد", "تنظیمات با موفقیت ذخیره شد.")
        else:
            QMessageBox.warning(self, "خطا", "خطا در ذخیره تنظیمات.")

    def ask_admin_password(self):
        pwd, ok = QInputDialog.getText(self, "رمز مدیر", "رمز عبور مدیر را وارد کنید:", QLineEdit.Password)
        return ok and pwd == ADMIN_PASSWORD

    def add_user(self):
        if not self.ask_admin_password(): return
        new_user = normalize_farsi(self.user_input.text())
        if not new_user or new_user in [self.user_list.item(i).text() for i in range(self.user_list.count())]: return
        self.add_user_item(new_user)
        self.user_input.clear()

    def add_user_item(self, username):
        self.user_list.addItem(QListWidgetItem(username))

    def remove_selected_user(self):
        if not self.user_list.selectedItems() or not self.ask_admin_password(): return
        for item in self.user_list.selectedItems():
            self.user_list.takeItem(self.user_list.row(item))

    def show_about(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("About")
        dlg.setFixedSize(500, 450)
        main_layout = QVBoxLayout(dlg)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)
        intro_layout = QVBoxLayout()
        intro_layout.setSpacing(0)
        intro_layout.setContentsMargins(0, 0, 0, 0)
        intro_text = "<h3><b>Fardan Apex — Serializer</b></h3><h4>Production Serial Generator</h4><br>Version: 6.7 — © 2025 All Rights Reserved<br>Developed for: Garma Gostar Fardan Co."
        lbl_intro = QLabel(intro_text)
        lbl_intro.setWordWrap(True)
        lbl_intro.setAlignment(Qt.AlignLeft)
        intro_layout.addWidget(lbl_intro)
        logo = QLabel()
        logo_pix = QPixmap(resource_path("FardanLogo.jpg")) or QPixmap(resource_path("FardanLogoEN.png"))
        if not logo_pix.isNull():
            logo.setPixmap(logo_pix.scaledToWidth(175, Qt.SmoothTransformation))
        logo.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        logo.setContentsMargins(35, 10, 0, 0)
        intro_layout.addWidget(logo)
        main_layout.addLayout(intro_layout)
        dev_layout = QVBoxLayout()
        dev_layout.setSpacing(0)
        dev_layout.setContentsMargins(5, 0, 0, 5)
        font_id = QFontDatabase.addApplicationFont(resource_path("BrittanySignature.ttf"))
        font_family = QFontDatabase.applicationFontFamilies(font_id)[0] if font_id != -1 else "Sans Serif"
        dev_text = f"<b>Design & Development:</b><br><span style='font-family:\"{font_family}\"; font-size:20pt; color:#4169E1;'>&nbsp;&nbsp;&nbsp;&nbsp;Behnam Rabieyan</span><br>website: behnamrabieyan.ir | E-mail: info@behnamrabieyan.ir"
        lbl_dev = QLabel(dev_text)
        lbl_dev.setWordWrap(True)
        lbl_dev.setAlignment(Qt.AlignLeft)
        dev_layout.addWidget(lbl_dev)
        main_layout.addLayout(dev_layout)
        dlg.exec_()

def show_splash(is_authorized: bool):
    app = QApplication.instance()
    splash_pix = QPixmap(resource_path("SerializerFardanApex.png"))
    splash = QSplashScreen(splash_pix, Qt.WindowStaysOnTopHint)
    splash.setMask(splash_pix.mask())
    progress = QProgressBar(splash)
    progress.setGeometry(90, splash_pix.height() - 100, splash_pix.width() - 180, 20)
    progress.setMaximum(100)
    progress.setStyleSheet("QProgressBar { border: 1px solid grey; border-radius: 5px; text-align: center; } QProgressBar::chunk { background-color: #2e7dff; }")
    splash.show()
    main_window = App(is_authorized=is_authorized)
    timer, step = QTimer(), 0
    def update_progress():
        nonlocal step
        step += 2
        progress.setValue(step)
        if step >= 100:
            timer.stop()
            splash.close()
            main_window.show()
    timer.timeout.connect(update_progress)
    timer.start(25)
    app.exec_()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    initialize_settings()
    current_user = getpass.getuser()
    is_authorized = not (ALLOWED_USERS and current_user not in ALLOWED_USERS)
    if not is_authorized:
        QMessageBox.warning(None, "حالت مهمان", f"کاربر '{current_user}' مجوز ویرایش ندارد.\nبرنامه بدون مجوز ثبت و ویرایش اجرا می‌شود.")
    QFontDatabase.addApplicationFont(resource_path("IRAN.ttf"))
    app.setFont(QFont("IRAN", 10))
    app.setWindowIcon(QIcon(resource_path("icon.ico")))
    show_splash(is_authorized=is_authorized)



